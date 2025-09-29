import os
import time
import pandas as pd
from selenium import webdriver
from openpyxl import load_workbook
from datetime import datetime, timedelta
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# --- HELPERS ---

def get_target_date():
    """
    Eğer bugün Pazartesi ise 3 gün önceyi (Cuma) döndür,
    diğer günlerde 1 gün önceyi döndür.
    """
    today = datetime.today()
    if today.weekday() == 0:  # Pazartesi
        return today - timedelta(days=3)
    else:
        return today - timedelta(days=1)

def setup_driver(download_dir):
    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "directory_upgrade": True
    }
    options.add_experimental_option("prefs", prefs)
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    return webdriver.Chrome(options=options)

def wait_for_file(download_dir, extensions=(".xls", ".xlsx"), timeout=30):
    for _ in range(timeout):
        files = [f for f in os.listdir(download_dir) if f.endswith(extensions)]
        if files:
            return max(files, key=lambda f: os.path.getctime(os.path.join(download_dir, f)))
        time.sleep(1)
    return None

def convert_xls_to_html(xls_path, html_path):
    wb = load_workbook(xls_path, data_only=True)
    ws = wb.active

    def format_cell_value(cell):
        val = cell.value
        if val is None:
            return ""
        if isinstance(val, (int, float)):
            fmt = cell.number_format
            try:
                # Format numeric values based on Excel number format
                if "0.00" in fmt or "0,00" in fmt:
                    s = f"{val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                    return s
                else:
                    s = f"{int(val):,}".replace(",", ".")
                    return s
            except:
                return str(val)
        else:
            return str(val)

    data = []
    for row in ws.iter_rows():
        row_data = [format_cell_value(cell) for cell in row]
        data.append(row_data)

    # Remove empty rows
    data = [row for row in data if any(cell != "" for cell in row)]

    # Remove empty columns
    if data:
        cols = len(data[0])
        non_empty_cols = [i for i in range(cols) if any(row[i] != "" for row in data)]
        data = [[row[i] for i in non_empty_cols] for row in data]

    df = pd.DataFrame(data[1:], columns=data[0])
    df.to_html(html_path, index=False, border=1, na_rep="")

# --- MAIN FUNCTION ---

def main():
    if datetime.today().weekday() >= 5:
        print("🛑 Today is weekend, script will not run.")
        return

    # Hedef tarih: Pazartesi ise 3 gün önce, değilse 1 gün önce
    target_date = get_target_date()
    date_str = target_date.strftime("%d/%m/%Y")
    print(f"📅 Target date for report: {date_str}")

    # Klasör oluştur
    data_dir = os.path.join(os.getcwd(), "data")
    os.makedirs(data_dir, exist_ok=True)

    # Webdriver başlat
    driver = setup_driver(data_dir)
    try:
        print("🌐 Opening webpage...")
        driver.get("https://www.vap.org.tr/api/all-companies")

        # Datepicker yüklenmesini bekle
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "input#datepicker")))

        # Tarihi gir ve değişikliği tetikle
        driver.execute_script(f"""
            var dateInput = document.querySelector("#datepicker");
            dateInput.value = '{date_str}';
            dateInput.dispatchEvent(new Event('change', {{ bubbles: true }}));
        """)

        # Gönder butonuna tıkla
        driver.find_element(By.CSS_SELECTOR, "input.submit-btn").click()
        print(f"📥 Downloading report for {date_str}...")

        # Dosya indirildi mi kontrol et
        downloaded_file = wait_for_file(data_dir)
        if not downloaded_file:
            print("❌ Download timed out.")
            return

    finally:
        driver.quit()

    xls_path = os.path.join(data_dir, downloaded_file)

    # HTML dosya yolları
    html_filename_date = os.path.splitext(downloaded_file)[0] + ".html"
    html_path_date = os.path.join(data_dir, html_filename_date)

    base_name = downloaded_file.split("-")[0]  # e.g. "Fiili_Dolasim_Raporu_MKK"
    html_filename_fixed = f"{base_name}.html"
    html_path_fixed = os.path.join(data_dir, html_filename_fixed)

    print(f"📄 Downloaded file: {downloaded_file}")
    print("🔄 Converting to cleaned HTML...")

    try:
        convert_xls_to_html(xls_path, html_path_date)
        convert_xls_to_html(xls_path, html_path_fixed)
        print(f"✅ HTML files created: {html_filename_date}, {html_filename_fixed}")
    except Exception as e:
        print(f"❌ Error during HTML conversion: {e}")

# --- ENTRY POINT ---

if __name__ == "__main__":
    main()
